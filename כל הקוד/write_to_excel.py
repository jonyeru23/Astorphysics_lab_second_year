import openpyxl
import pandas as pd
from SFR import *
from get_SFR_graphs import *

def testing():
    file_path = "../testing.xlsx"
    wb = openpyxl.load_workbook(file_path)
    ws = wb["Sheet1"]
    ws['A2'] = 555
    wb.save(file_path)


def write_Luminosity_lambda():
    file_path = "../galaxies.xlsx"
    df = pd.read_excel(file_path, sheet_name="summary")
    wb = openpyxl.load_workbook(file_path)
    for i, band in enumerate(["b", "i", "u", "v"]):
        y, yerr = clean(*get_L_lambda(df, band))
        for galaxy in range(34):
            ws = wb[str(galaxy+1)]
            # set title
            ws['AU1'] = "L_lambda[erg/s/A]"
            ws['AV1'] = "dL_lambda[erg/s/A]"

            # set data
            ws[f"AU{i+2}"] = y[galaxy]
            ws[f"AV{i+2}"] = yerr[galaxy]

    wb.save(file_path)




def write_to_same_cell_SFR():
    file_path = "../galaxies.xlsx"
    df = pd.read_excel(file_path, sheet_name="summary")
    print(df)
    y, yerr = clean(*get_SFR(df))
    print(len(y))
    wb = openpyxl.load_workbook(file_path)

    for i in range(34):
        ws = wb[str(i+1)]
        ws['AS1'] = "SFR[solMass/year]"
        ws['AT1'] = "dSFR[solMass/year]"

        ws['AS2'] = y[i]
        ws['AT2'] = yerr[i]

    wb.save(file_path)


def move_to_rest_frame():
    filters = ["b", "i", "u", "v"]
    file_path = "../galaxies.xlsx"
    df = pd.read_excel(file_path, sheet_name="summary")
    wb = openpyxl.load_workbook(file_path)
    ws = wb["summary"]
    rest_frames = {
        "b": df["pick b"],
        "i": df["pick i"],
        "u": df["pick u"],
        "v": df["pick v"]
    }

    for cul, filter_ in zip(["BH", "BJ", "BL", "BN"], filters):
        for i in range(34):
            ws[f"{cul}{i+2}"] = df[f"dFlux {rest_frames[filter_][i]}"][i]

    wb.save(file_path)






if __name__ == '__main__':
    # write_Luminosity_lambda()
    # write_to_same_cell_SFR()
    move_to_rest_frame()